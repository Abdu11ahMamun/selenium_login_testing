import time
import unittest
from selenium import webdriver
from openpyxl import load_workbook
import threading

class UserLoginThread(threading.Thread):
    def __init__(self, username, password):
        super().__init__()
        self.username = username
        self.password = password

    def run(self):
        browser = webdriver.Chrome()  # Change to Chrome driver
        browser.get('http://192.168.1.154:8080/beftn/faces/login.xhtml')

        # Delay for 10 seconds
        #time.sleep(10)
 
        # Fill in the login form
        username_input = browser.find_element('xpath', '//*[@id="j_idt13"]/table[2]/tbody/tr[2]/td/input')
        password_input = browser.find_element('xpath', '//*[@id="j_idt13:masked-password"]')
        submit_button = browser.find_element('xpath', '//*[@id="j_idt13"]/table[2]/tfoot/tr/td/input')

        username_input.send_keys(self.username)
        password_input.send_keys(self.password)
        submit_button.click()

        # Verify the URL after login
        expected_url = 'http://192.168.1.154:8080/beftn/faces/index.xhtml'
        if browser.current_url == expected_url:
            print(f"Success: Login Successful! User: {self.username}")
        else:
            print(f"Failure: Login Failed! User: {self.username}")

        browser.quit()

class LoadTest(unittest.TestCase):
    def test_user_load(self):
        # Load usernames and passwords from Excel sheet
        workbook = load_workbook('D:/BEFTN/userinforbook.xlsx')
        sheet = workbook.active

        threads = []
        num_users = 100

        # Create and start user login threads
        for i in range(1, num_users+1):
            username = sheet.cell(row=i, column=1).value
            password = sheet.cell(row=i, column=2).value
            thread = UserLoginThread(username, password)
            threads.append(thread)
            thread.start()

        # Wait for all threads to complete
        for thread in threads:
            thread.join()

if __name__ == '__main__':
    unittest.main(verbosity=2)
